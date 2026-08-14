"""Read TARGET-AML open clinical supplement workbooks without concatenating them."""

from __future__ import annotations

from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pediastat.ingestion.identifiers import (
    identifier_shape,
    join_barcode,
    normalize_identifier,
    original_identifier,
    summarize_identifiers,
)
from pediastat.ingestion.missingness import classify_missing, is_observed

IDENTIFIER_HEADERS = frozenset({"target usi", "target barcode", "submitter_id"})


def cell_to_jsonable(value: Any) -> Any:
    """Convert an Excel cell value to a JSON-serializable form."""
    if value is None:
        return None
    if isinstance(value, datetime | date):
        return value.isoformat()
    if isinstance(value, str):
        return value
    if isinstance(value, bool | int):
        return value
    if isinstance(value, float):
        return value
    return str(value)


def _header_names(row: tuple[Any, ...] | None) -> list[str]:
    if row is None:
        return []
    names: list[str] = []
    for index, value in enumerate(row):
        if value is None or str(value).strip() == "":
            names.append(f"unnamed_{index + 1}")
        else:
            names.append(str(value).strip())
    return names


def identifier_header(columns: list[str]) -> str | None:
    for name in columns:
        if name.strip().lower() in IDENTIFIER_HEADERS:
            return name
    return None


def infer_primitive_type(values: list[Any]) -> str:
    observed = [value for value in values if is_observed(value)]
    if not observed:
        return "empty"
    kinds: set[str] = set()
    for value in observed:
        if isinstance(value, bool):
            kinds.add("bool")
        elif isinstance(value, int):
            kinds.add("int")
        elif isinstance(value, float):
            kinds.add("float")
        else:
            kinds.add("str")
    if len(kinds) == 1:
        return next(iter(kinds))
    if kinds <= {"int", "float"}:
        return "float"
    return "mixed"


def read_workbook_sheets(path: Path) -> list[dict[str, Any]]:
    """Read every sheet in a workbook, preserving original column names."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: list[dict[str, Any]] = []
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows_iter = worksheet.iter_rows(values_only=True)
            header = next(rows_iter, None)
            columns = _header_names(header)
            id_col = identifier_header(columns)
            records: list[dict[str, Any]] = []
            identifiers: list[Any] = []
            for row_number, row in enumerate(rows_iter, start=2):
                cells: dict[str, Any] = {}
                for index, column in enumerate(columns):
                    raw = row[index] if index < len(row) else None
                    cells[column] = cell_to_jsonable(raw)
                identifier = cells.get(id_col) if id_col else None
                identifiers.append(identifier)
                records.append(
                    {
                        "row_number": row_number,
                        "original_identifier": original_identifier(identifier),
                        "normalized_identifier": normalize_identifier(identifier),
                        "join_barcode": join_barcode(identifier),
                        "identifier_shape": identifier_shape(identifier),
                        "cells": cells,
                    }
                )
            sheets.append(
                {
                    "workbook": path.name,
                    "path": str(path),
                    "sheet": sheet_name,
                    "columns": columns,
                    "identifier_field": id_col,
                    "n_rows": len(records),
                    "n_columns": len(columns),
                    "is_patient_level": id_col is not None,
                    "records": records,
                    "identifier_summary": summarize_identifiers(identifiers),
                }
            )
    finally:
        workbook.close()
    return sheets


def profile_sheet(
    sheet: dict[str, Any], *, include_samples: bool = False
) -> list[dict[str, Any]]:
    """Build column profiles. Sample values are omitted unless requested."""
    columns: list[str] = sheet["columns"]
    records: list[dict[str, Any]] = sheet["records"]
    n_rows = len(records)
    profiles: list[dict[str, Any]] = []
    id_field = sheet.get("identifier_field")
    for column in columns:
        values = [record["cells"].get(column) for record in records]
        missing = sum(classify_missing(value) != "observed" for value in values)
        observed = [value for value in values if is_observed(value)]
        unique = {repr(value) for value in observed}
        inferred = infer_primitive_type(values)
        numeric = [
            float(value)
            for value in observed
            if isinstance(value, int | float) and not isinstance(value, bool)
        ]
        sample: list[str] | None = None
        if include_samples and column != id_field:
            sample = [str(value) for value in observed[:5]]
        profiles.append(
            {
                "workbook": sheet["workbook"],
                "sheet": sheet["sheet"],
                "column": column,
                "is_identifier": column == id_field,
                "n_rows": n_rows,
                "n_missing": missing,
                "pct_missing": round(missing / n_rows * 100.0, 2) if n_rows else 0.0,
                "n_unique_observed": len(unique),
                "inferred_type": inferred,
                "numeric_min": min(numeric) if numeric else None,
                "numeric_max": max(numeric) if numeric else None,
                "sample_values": sample,
            }
        )
    return profiles


def column_value_counts(sheet: dict[str, Any], column: str) -> Counter[str]:
    counts: Counter[str] = Counter()
    for record in sheet["records"]:
        value = record["cells"].get(column)
        if is_observed(value):
            counts[str(value)] += 1
    return counts
