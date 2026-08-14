"""Orchestrate the TARGET-AML GDC source audit and write metadata artifacts."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from collections.abc import Mapping, Sequence
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from pediastat.audit.client import GDCAPIError, GDCClient
from pediastat.audit.constants import (
    CANDIDATE_FIELDS,
    CLINICAL_FILE_FIELDS,
    DEFAULT_PAGE_SIZE,
    DEFAULT_TIMEOUT_SECONDS,
    GDC_API_BASE_URL,
    TARGET_AML_PROJECT_ID,
)
from pediastat.audit.extract import field_exists_in_mapping, mapping_field_names
from pediastat.audit.filters import (
    target_aml_cases_filter,
    target_aml_clinical_files_filter,
)
from pediastat.audit.summarize import entity_cardinality, summarize_field
from pediastat.audit.survival import audit_survival_fields
from pediastat.config import PROJECT_ROOT

DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "artifacts" / "source_audit"
DEFAULT_DOWNLOAD_DIR = PROJECT_ROOT / "data" / "raw" / "gdc_open_clinical_supplements"


def utc_now_iso() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat()


def fetch_project(
    client: GDCClient, project_id: str = TARGET_AML_PROJECT_ID
) -> dict[str, Any]:
    payload = client.get_json(
        f"projects/{project_id}",
        params={
            "expand": "summary,summary.experimental_strategies,summary.data_categories",
        },
    )
    data = payload.get("data")
    if not isinstance(data, dict):
        raise GDCAPIError(f"Unexpected project payload for {project_id}")
    return data


def fetch_cases_mapping(client: GDCClient) -> dict[str, Any]:
    payload = client.get_json("cases/_mapping")
    if not isinstance(payload, dict):
        raise GDCAPIError("Unexpected cases/_mapping payload")
    return payload


def fetch_all_cases(
    client: GDCClient,
    fields: Sequence[str],
    page_size: int = DEFAULT_PAGE_SIZE,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    hits: list[dict[str, Any]] = []
    pagination: dict[str, Any] = {}
    from_index = 0
    while True:
        payload = client.post_json(
            "cases",
            {
                "filters": target_aml_cases_filter(),
                "fields": ",".join(fields),
                "size": page_size,
                "from": from_index,
            },
        )
        data = payload.get("data", {})
        page_hits = data.get("hits", [])
        pagination = data.get("pagination", {})
        if not isinstance(page_hits, list):
            raise GDCAPIError("Unexpected cases payload: hits is not a list")
        hits.extend(item for item in page_hits if isinstance(item, dict))
        total = int(pagination.get("total") or 0)
        from_index += len(page_hits)
        if from_index >= total or not page_hits:
            break
    return hits, pagination


def fetch_clinical_files(client: GDCClient) -> list[dict[str, Any]]:
    payload = client.post_json(
        "files",
        {
            "filters": target_aml_clinical_files_filter(),
            "fields": ",".join(CLINICAL_FILE_FIELDS),
            "size": 100,
        },
    )
    hits = payload.get("data", {}).get("hits", [])
    if not isinstance(hits, list):
        raise GDCAPIError("Unexpected files payload: hits is not a list")
    return [item for item in hits if isinstance(item, dict)]


def _project_ids_from_file(file_hit: Mapping[str, Any]) -> list[str]:
    projects: list[str] = []
    for case in file_hit.get("cases") or []:
        if not isinstance(case, dict):
            continue
        project = case.get("project")
        if isinstance(project, dict) and project.get("project_id"):
            project_id = str(project["project_id"])
            if project_id not in projects:
                projects.append(project_id)
    return projects


def flatten_clinical_file_row(file_hit: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "file_id": file_hit.get("file_id") or file_hit.get("id"),
        "file_name": file_hit.get("file_name"),
        "data_category": file_hit.get("data_category"),
        "data_type": file_hit.get("data_type"),
        "data_format": file_hit.get("data_format"),
        "access": file_hit.get("access"),
        "file_size": file_hit.get("file_size"),
        "md5sum": file_hit.get("md5sum"),
        "state": file_hit.get("state"),
        "associated_project": ";".join(_project_ids_from_file(file_hit)),
        "n_associated_cases": len(file_hit.get("cases") or []),
    }


def md5_file(path: Path) -> str:
    digest = hashlib.md5(usedforsecurity=False)
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 64), b""):
            digest.update(chunk)
    return digest.hexdigest()


def inspect_xlsx_columns(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: list[dict[str, Any]] = []
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = worksheet.iter_rows(values_only=True)
            header_row = next(rows, None)
            columns = [
                str(value) if value is not None else ""
                for value in (header_row or ())
            ]
            n_data_rows = sum(1 for _ in rows)
            sheets.append(
                {
                    "sheet": sheet_name,
                    "n_columns": len(columns),
                    "n_data_rows": n_data_rows,
                    "columns": columns,
                }
            )
    finally:
        workbook.close()
    return sheets


def download_open_supplements(
    client: GDCClient,
    file_rows: Sequence[Mapping[str, Any]],
    download_dir: Path,
) -> list[dict[str, Any]]:
    inspections: list[dict[str, Any]] = []
    for row in file_rows:
        access = str(row.get("access") or "").lower()
        file_id = str(row.get("file_id") or "")
        file_name = str(row.get("file_name") or file_id)
        record: dict[str, Any] = {
            "file_id": file_id,
            "file_name": file_name,
            "access": row.get("access"),
            "source_url": f"{client.base_url}/data/{file_id}",
            "api_endpoint": f"{client.base_url}/files",
        }
        if access != "open":
            record["downloaded"] = False
            record["skip_reason"] = "not open-access; download refused"
            inspections.append(record)
            continue
        destination = download_dir / file_name
        downloaded_at = utc_now_iso()
        client.download_open_file(file_id, destination)
        checksum = md5_file(destination)
        record.update(
            {
                "downloaded": True,
                "download_timestamp_utc": downloaded_at,
                "local_path": str(destination.relative_to(PROJECT_ROOT)),
                "checksum_md5_downloaded": checksum,
                "checksum_md5_api": row.get("md5sum"),
                "checksum_match": (
                    checksum == row.get("md5sum") if row.get("md5sum") else None
                ),
            }
        )
        if destination.suffix.lower() == ".xlsx":
            record["sheets"] = inspect_xlsx_columns(destination)
        inspections.append(record)
    return inspections


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(payload, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _write_field_csv(path: Path, rows: Sequence[Mapping[str, Any]]) -> None:
    fieldnames = [
        "field_path",
        "source_entity",
        "exists_in_mapping",
        "gdc_type",
        "gdc_description",
        "n_cases",
        "n_available",
        "n_missing",
        "pct_missing",
        "n_usable",
        "value_kind",
        "numeric_min",
        "numeric_max",
        "n_negative_values",
        "n_zero_values",
        "n_distinct",
        "distinct_values",
        "n_cases_with_multiple_nonnull",
        "n_cases_with_disagreeing_values",
        "n_nested_records_observed",
        "n_gdc_missing_code_records",
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            serialized = dict(row)
            distinct = serialized.get("distinct_values")
            if isinstance(distinct, list):
                serialized["distinct_values"] = " | ".join(
                    str(item) for item in distinct
                )
            writer.writerow(serialized)


def _write_files_csv(path: Path, rows: Sequence[Mapping[str, Any]]) -> None:
    fieldnames = [
        "file_id",
        "file_name",
        "data_category",
        "data_type",
        "data_format",
        "access",
        "file_size",
        "md5sum",
        "state",
        "associated_project",
        "n_associated_cases",
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({name: row.get(name) for name in fieldnames})


def run_audit(
    *,
    client: GDCClient | None = None,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    download_dir: Path = DEFAULT_DOWNLOAD_DIR,
    skip_download: bool = False,
    page_size: int = DEFAULT_PAGE_SIZE,
) -> dict[str, Any]:
    """Run the source audit and write metadata artifacts.

    Clinical case records are not written to disk.
    """
    client = client or GDCClient(timeout_seconds=max(DEFAULT_TIMEOUT_SECONDS, 120.0))
    audit_started = utc_now_iso()
    project = fetch_project(client)
    mapping_payload = fetch_cases_mapping(client)
    indexed_mapping = mapping_field_names(mapping_payload)

    requested_fields = [
        field
        for field in CANDIDATE_FIELDS
        if field_exists_in_mapping(indexed_mapping, field)
    ]
    missing_from_mapping = [
        field
        for field in CANDIDATE_FIELDS
        if not field_exists_in_mapping(indexed_mapping, field)
    ]

    cases, pagination = fetch_all_cases(client, requested_fields, page_size=page_size)
    field_summaries = []
    for field in CANDIDATE_FIELDS:
        exists = field_exists_in_mapping(indexed_mapping, field)
        meta = indexed_mapping.get(field) or indexed_mapping.get(f"cases.{field}") or {}
        summary = summarize_field(
            cases,
            field,
            exists_in_mapping=exists,
            gdc_type=meta.get("type") if exists else None,
            gdc_description=meta.get("description") if exists else None,
        )
        field_summaries.append(summary)

    survival = audit_survival_fields(cases)
    cardinality = entity_cardinality(cases)
    file_hits = fetch_clinical_files(client)
    file_rows = [flatten_clinical_file_row(item) for item in file_hits]

    supplement_inspections: list[dict[str, Any]] = []
    if not skip_download:
        supplement_inspections = download_open_supplements(
            client, file_rows, download_dir
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    project_metadata = {
        "audit_timestamp_utc": audit_started,
        "api_base_url": client.base_url,
        "project_id": project.get("project_id"),
        "name": project.get("name"),
        "disease_type": project.get("disease_type"),
        "primary_site": project.get("primary_site"),
        "dbgap_accession_number": project.get("dbgap_accession_number"),
        "releasable": project.get("releasable"),
        "released": project.get("released"),
        "state": project.get("state"),
        "summary": project.get("summary"),
        "cases_api_pagination_total": pagination.get("total"),
        "n_cases_retrieved": len(cases),
        "fields_missing_from_current_mapping": missing_from_mapping,
        "note": (
            "Case counts and field availability are specific to this GDC release "
            "and audit timestamp. Do not treat them as permanent."
        ),
    }
    _write_json(output_dir / "project_metadata.json", project_metadata)
    _write_field_csv(output_dir / "clinical_field_availability.csv", field_summaries)
    _write_json(output_dir / "survival_field_audit.json", survival)
    _write_json(output_dir / "entity_cardinality.json", cardinality)
    _write_files_csv(output_dir / "open_clinical_files.csv", file_rows)
    if supplement_inspections:
        _write_json(
            output_dir / "open_clinical_supplement_columns.json",
            supplement_inspections,
        )

    return {
        "project_metadata": project_metadata,
        "field_summaries": field_summaries,
        "survival": survival,
        "cardinality": cardinality,
        "clinical_files": file_rows,
        "supplement_inspections": supplement_inspections,
        "n_cases": len(cases),
        "output_dir": str(output_dir),
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Audit public TARGET-AML clinical metadata from the GDC API."
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Directory for metadata summaries (not a clinical dataset).",
    )
    parser.add_argument(
        "--download-dir",
        type=Path,
        default=DEFAULT_DOWNLOAD_DIR,
        help="Directory for open-access clinical supplement files.",
    )
    parser.add_argument(
        "--skip-download",
        action="store_true",
        help="Do not download clinical supplement files.",
    )
    parser.add_argument(
        "--page-size",
        type=int,
        default=DEFAULT_PAGE_SIZE,
    )
    args = parser.parse_args(argv)

    result = run_audit(
        output_dir=args.output_dir,
        download_dir=args.download_dir,
        skip_download=args.skip_download,
        page_size=args.page_size,
    )
    survival = result["survival"]
    print("TARGET-AML source audit")
    print(f"API: {GDC_API_BASE_URL}")
    print(f"Cases retrieved: {result['n_cases']}")
    print(
        "Possible GDC-style OS times: "
        f"{survival['n_cases_with_possible_gdc_style_survival_time']}"
    )
    print(f"Artifacts: {result['output_dir']}")
    print("No clinical case-level dataset was written.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
